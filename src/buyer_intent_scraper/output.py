"""CSV and Excel (.xlsx) output."""

from __future__ import annotations

import csv
from pathlib import Path

from buyer_intent_scraper.models import Lead

CSV_FIELDS = [
    "name",
    "entity_type",
    "service",
    "location",
    "intent_signal",
    "emails",
    "phones",
    "website",
    "source_type",
    "source_title",
    "source_url",
    "published_date",
    "confidence",
    "discovered_at",
]


def lead_to_row(lead: Lead) -> dict[str, str]:
    return {
        "name": lead.name,
        "entity_type": lead.entity_type,
        "service": lead.service,
        "location": lead.location,
        "intent_signal": lead.intent_signal,
        "emails": "; ".join(lead.emails),
        "phones": "; ".join(lead.phones),
        "website": lead.website,
        "source_type": lead.source_type,
        "source_title": lead.source_title,
        "source_url": lead.source_url,
        "published_date": lead.published_date,
        "confidence": f"{lead.confidence:.3f}",
        "discovered_at": lead.discovered_at,
    }


def write_csv(leads: list[Lead], path: str | Path) -> Path:
    out_path = Path(path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=CSV_FIELDS)
        writer.writeheader()
        for lead in leads:
            writer.writerow(lead_to_row(lead))
    return out_path


# Per-column display widths for the Excel sheet.
_XLSX_WIDTHS = {
    "name": 32,
    "entity_type": 13,
    "service": 22,
    "location": 18,
    "intent_signal": 60,
    "emails": 30,
    "phones": 22,
    "website": 24,
    "source_type": 14,
    "source_title": 36,
    "source_url": 40,
    "published_date": 14,
    "confidence": 11,
    "discovered_at": 22,
}


def write_xlsx(leads: list[Lead], path: str | Path, sheet_name: str = "Leads") -> Path:
    """Write leads to a styled .xlsx workbook with clickable links + auto-filter."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    out_path = Path(path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31] or "Leads"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    link_font = Font(color="0563C1", underline="single")

    ws.append(CSV_FIELDS)
    for col_idx, field in enumerate(CSV_FIELDS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = _XLSX_WIDTHS.get(field, 18)

    url_cols = {CSV_FIELDS.index("source_url") + 1, CSV_FIELDS.index("website") + 1}
    conf_col = CSV_FIELDS.index("confidence") + 1

    for lead in leads:
        row = lead_to_row(lead)
        ws.append([row[field] for field in CSV_FIELDS])
        r = ws.max_row
        for col_idx in url_cols:
            cell = ws.cell(row=r, column=col_idx)
            value = str(cell.value or "")
            if value:
                target = value if value.startswith("http") else f"https://{value}"
                cell.hyperlink = target
                cell.font = link_font
        conf_cell = ws.cell(row=r, column=conf_col)
        try:
            conf_cell.value = float(conf_cell.value)
            conf_cell.number_format = "0.000"
        except (TypeError, ValueError):
            pass
        ws.cell(row=r, column=CSV_FIELDS.index("intent_signal") + 1).alignment = Alignment(
            wrap_text=True, vertical="top"
        )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(CSV_FIELDS))}{ws.max_row}"
    wb.save(out_path)
    return out_path


def write_output(leads: list[Lead], path: str | Path, fmt: str = "xlsx") -> list[Path]:
    """Write leads in the requested format(s).

    ``fmt`` is one of ``"xlsx"`` (default), ``"csv"`` or ``"both"``. ``path`` may
    carry an extension; it is normalized per format.
    """
    base = Path(path)
    base = base.with_suffix("")
    written: list[Path] = []
    if fmt in ("xlsx", "both"):
        written.append(write_xlsx(leads, base.with_suffix(".xlsx")))
    if fmt in ("csv", "both"):
        written.append(write_csv(leads, base.with_suffix(".csv")))
    if not written:
        raise ValueError(f"Unknown output format: {fmt!r} (use xlsx, csv or both)")
    return written
