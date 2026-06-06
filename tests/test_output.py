import csv

from openpyxl import load_workbook

from buyer_intent_scraper.models import Lead
from buyer_intent_scraper.output import CSV_FIELDS, write_csv, write_output, write_xlsx


def _sample_leads():
    return [
        Lead(
            name="Acme Builders Ltd",
            service="construction services",
            location="Nairobi, Kenya",
            intent_signal="Looking for a contractor",
            source_type="google_dork",
            source_url="https://acme.co.ke",
            emails=["info@acme.co.ke"],
            phones=["+254712345678"],
            website="acme.co.ke",
            confidence=0.85,
        )
    ]


def test_write_csv(tmp_path):
    leads = [
        Lead(
            name="Acme Builders Ltd",
            service="construction services",
            location="Nairobi, Kenya",
            intent_signal="Looking for a contractor",
            source_type="google_dork",
            source_url="https://acme.co.ke",
            emails=["info@acme.co.ke", "procurement@acme.co.ke"],
            phones=["+254712345678"],
            website="acme.co.ke",
            confidence=0.85,
        )
    ]
    out = tmp_path / "leads.csv"
    write_csv(leads, out)

    with out.open(newline="", encoding="utf-8") as fh:
        rows = list(csv.DictReader(fh))

    assert len(rows) == 1
    row = rows[0]
    assert list(row.keys()) == CSV_FIELDS
    assert row["name"] == "Acme Builders Ltd"
    assert row["emails"] == "info@acme.co.ke; procurement@acme.co.ke"
    assert row["phones"] == "+254712345678"
    assert row["confidence"] == "0.850"


def test_write_csv_creates_dirs(tmp_path):
    out = tmp_path / "nested" / "dir" / "leads.csv"
    write_csv([], out)
    assert out.exists()


def test_write_xlsx(tmp_path):
    out = tmp_path / "leads.xlsx"
    write_xlsx(_sample_leads(), out)
    assert out.exists()

    wb = load_workbook(out)
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert header == CSV_FIELDS
    row = [c.value for c in ws[2]]
    assert row[CSV_FIELDS.index("name")] == "Acme Builders Ltd"
    assert row[CSV_FIELDS.index("emails")] == "info@acme.co.ke"
    # confidence is written as a real number, not text
    assert row[CSV_FIELDS.index("confidence")] == 0.85
    # source_url cell carries a clickable hyperlink
    url_cell = ws.cell(row=2, column=CSV_FIELDS.index("source_url") + 1)
    assert url_cell.hyperlink is not None
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref is not None


def test_write_output_formats(tmp_path):
    leads = _sample_leads()

    xlsx_paths = write_output(leads, tmp_path / "report", fmt="xlsx")
    assert xlsx_paths == [tmp_path / "report.xlsx"]
    assert xlsx_paths[0].exists()

    both_paths = write_output(leads, tmp_path / "report2.csv", fmt="both")
    suffixes = sorted(p.suffix for p in both_paths)
    assert suffixes == [".csv", ".xlsx"]
    assert all(p.exists() for p in both_paths)
